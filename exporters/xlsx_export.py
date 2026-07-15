"""
exporters/xlsx_export.py
---------------------------
Excel (.xlsx) exporter using openpyxl — a structured, filterable deal
sheet plus a summary tab, which is what a business user would actually
want to pivot/filter on.
"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_xlsx(newsletter: Dict[str, Any], out_path: str) -> str:
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: Deals ---
    ws = wb.active
    ws.title = "Deals"
    headers = [
        "Title", "Published", "Deal Type", "Acquirer", "Target", "Amount", "Sector",
        "Country", "Credibility", "Credibility Score", "Relevance Score",
        "Source Domain", "Link",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for d in newsletter.get("deals", []):
        ws.append([
            d.get("title", ""),
            d.get("published", "Date unknown"),
            d.get("deal_type", "").replace("_", " ").title(),
            d.get("acquirer") or "—",
            d.get("target") or "—",
            d.get("amount") or "—",
            d.get("sector", "FMCG"),
            d.get("country") or "—",
            d.get("credibility_label", ""),
            d.get("credibility_score", 0),
            d.get("relevance_score", 0),
            d.get("source", ""),
            d.get("link", ""),
        ])

    widths = [45, 14, 16, 20, 20, 14, 14, 12, 30, 12, 12, 22, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Generated At", newsletter.get("generated_at", "")])
    ws2.append([])
    ws2.append(["Executive Summary"])
    ws2.append([newsletter.get("executive_summary", "")])
    ws2.append([])
    ws2.append(["Deal Type", "Count"])
    for dtype, count in sorted(newsletter.get("trend_snapshot", {}).items(), key=lambda x: -x[1]):
        ws2.append([dtype.replace("_", " ").title(), count])
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 60

    wb.save(out_path)
    return out_path
